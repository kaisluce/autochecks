import os
import sys
from typing import Optional
import pandas as pd
from openpyxl import load_workbook
import xlsxwriter  # explicit import so PyInstaller bundles the Excel writer engine
import forSirenSiret.requestsiren as rsn
import forSirenSiret.requestsiret as rst
from thefuzz import fuzz

# Main SIREN/SIRET verification pipeline: builds Excel report with conditional formatting.

# Number of rows covered by conditional formatting rules (large enough for typical runs).
MAX_FORMAT_ROWS = 200000
# Définit la structure (ordre des colonnes) du rapport Excel final.
# Ceci est crucial pour la cohérence lors de l'écriture et de la relecture des données.
REPORT_COLUMNS = [
    "BP",
    "type",
    "siret",
    "nic",
    "siren",
    "status",
    "siege",
    "denomination",
    "Name 1",
    "date_creation",
    "date_cessation",
    "siret_siege",
    "naf",
    "naf_label",
    "cat_juridique",
    "adresse",
    "n_voie",
    "voie",
    "code_postal",
    "commune",
    "duplicates_siren",
    "duplicates_siret",
    "duplicates_VAT",
    "missing siren",
    "missing siret",
    "Missing_Vat",
    "uses a snetor siren",
    "uses a snetor siret",
    "uses a snetor VAT",
    "Missmatching siren siret",
    "Missmatching siren VAT",
    "Country/Region Key",
    "Language Key",
    "datas",
    "report",
    "diagnostic_name",
]
# Seuils pour la comparaison floue des noms et des rues.
# Un score supérieur à ce seuil est considéré comme une "légère différence".
NAME_SLIGHT_THRESHOLD = 75
STREET_SLIGHT_THRESHOLD = 75

ID_COLUMNS = ("BP", "Business Partner", "siren", "siret")

def _coerce_id_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Ensure identifier columns remain strings when reading external files."""
    for col in ID_COLUMNS:
        if col in df.columns:
            df[col] = df[col].astype(str)
    return df

def compare_names(row):
    """
    Compare le nom de l'entreprise provenant de l'API ('denomination') avec le nom
    provenant du fichier d'informations du partenaire ('Name 1') en utilisant la correspondance floue.

    Args:
        row (pd.Series): Une ligne du DataFrame contenant les colonnes 'denomination' et 'Name 1'.

    Returns:
        str: 'exact' pour une correspondance parfaite, 'slight difference' pour une correspondance proche,
             'no match' pour une faible correspondance, ou 'Missing name' si l'un des noms est manquant.
    """
    name1 = row.get('denomination')
    name2 = row.get('Name 1')
    if pd.isna(name1) or pd.isna(name2) or name1 == 'nan' or name2 == 'nan':
        return 'Missing name'
    
    score = fuzz.token_set_ratio(name1, name2)
    if score == 100:
        return 'exact'
    elif score > NAME_SLIGHT_THRESHOLD:
        return 'slight difference'
    else:
        return 'no match'

def _build_template(report_file: str, existing: Optional[pd.DataFrame] = None):
    """
    Crée un nouveau fichier de rapport Excel avec une feuille "Report", des en-têtes de colonnes
    et des règles de formatage conditionnel prédéfinies pour la coloration des lignes.

    Args:
        report_file (str): Le chemin où le fichier Excel sera créé.
        existing (Optional[pd.DataFrame]): Un DataFrame existant à écrire dans le
                                            modèle après sa création. Utilisé pour reconstruire
                                            des fichiers avec des données existantes.
    """
    with pd.ExcelWriter(report_file, engine="xlsxwriter", mode = "w") as writer:
        # Crée une feuille vide avec les en-têtes de colonnes corrects.
        pd.DataFrame(columns=REPORT_COLUMNS).to_excel(
            writer, index=False, sheet_name="Report", header=True
        )

        workbook = writer.book
        worksheet = writer.sheets["Report"]

        # --- Définitions des formats de couleur ---
        green = workbook.add_format({"bg_color": "#C6EFCE"})   # OK
        orange = workbook.add_format({"bg_color": "#FAB370"})  # missing / mismatch
        red = workbook.add_format({"bg_color": "#F4CCCC"})     # not active / invalid
        yellow = workbook.add_format({"bg_color": "#FAE484"})  # active + all good + duplicates
        dark_orange = workbook.add_format({"bg_color": "#BB9255"})

        # --- Règles de formatage conditionnel ---
        # Ces règles colorent les lignes en fonction de combinaisons de statuts, de validité et de doublons.
        # La plage est définie de manière très large pour couvrir la plupart des cas d'utilisation.
        range_ref = f"A2:AJ{MAX_FORMAT_ROWS}"

        # Flags (colonnes W -> AD) doivent tous être FALSE pour considérer la ligne "clean".
        # Flags columns X -> AE must all be FALSE to consider the line clean.
        flags_all_false = (
            "AND("
            '$X2="False",'
            '$Y2="False",'
            '$Z2="False",'
            '$AA2="False",'
            '$AB2="False",'
            '$AC2="False",'
            '$AD2="False",'
            '$AE2="False")'
        )
        flags_any_true = f"NOT({flags_all_false})"

        # Jaune : actif, aucun flag, mais des doublons.
        worksheet.conditional_format(
            range_ref,
            {
                "type": "formula",
                "criteria": (
                    '=AND('
                    'OR($F2="Actif",$F2="Active"),'
                    f'{flags_all_false},'
                    'OR($U2<>"[]",$V2<>"[]",$W2<>"[]")'
                    ')'
                ),
                "format": yellow,
                "stop_if_true": True,
            },
        )

        # Orange : actif, au moins un flag, doublons vides.
        worksheet.conditional_format(
            range_ref,
            {
                "type": "formula",
                "criteria": (
                    '=AND('
                    'OR($F2="Actif",$F2="Active"),'
                    f'{flags_any_true},'
                    '$U2="[]",'
                    '$V2="[]",'
                    '$W2="[]"'
                    ')'
                ),
                "format": orange,
                "stop_if_true": True,
            },
        )

        # Marron (dark_orange) : actif avec flags levés (peu importe les doublons).
        worksheet.conditional_format(
            range_ref,
            {
                "type": "formula",
                "criteria": (
                    "=AND("
                    'OR($F2="Actif",$F2="Active"),'
                    f"{flags_any_true}"
                    ")"
                ),
                "format": dark_orange,
                "stop_if_true": True,
            },
        )

        # Rouge : non actif.
        worksheet.conditional_format(
            range_ref,
            {
                "type": "formula",
                "criteria": '=NOT(OR($F2="Actif",$F2="Active"))',
                "format": red,
            },
        )

        # Vert : actif, aucun flag, aucun doublon.
        worksheet.conditional_format(
            range_ref,
            {
                "type": "formula",
                "criteria": (
                    '=AND('
                    'OR($F2="Actif",$F2="Active"),'
                    '$V2="[]",'
                    '$X2="False",'
                    '$Y2="False",'
                    '$Z2="False",'
                    '$AA2="False",'
                    '$AB2="False",'
                    '$AC2="False",'
                    '$AD2="False",'
                    '$AE2="False"'
                    ')'
                ),
                "format": green,
            },
        )

        # Si des données existantes sont fournies, les écrire dans le nouveau modèle.
        if existing is not None and not existing.empty:
            existing = existing.astype(str).reindex(columns=REPORT_COLUMNS)
            existing.to_excel(
                writer,
                index=False,
                header=False,
                sheet_name="Report",
                startrow=1,
            )

# Variante robuste : reconstruit si la feuille/les colonnes ne correspondent pas au gabarit.
def _ensure_template_v2(report_file: str):
    if not os.path.exists(report_file):
        _build_template(report_file)
        return
    try:
        wb = load_workbook(report_file)
    except Exception:
        _build_template(report_file)
        return

    needs_rebuild = "Report" not in wb.sheetnames
    try:
        existing = pd.read_excel(report_file, sheet_name="Report", dtype=str)
        existing = _coerce_id_columns(existing)
    except Exception:
        existing = pd.DataFrame()
        needs_rebuild = True

    if not needs_rebuild and list(existing.columns) != REPORT_COLUMNS:
        needs_rebuild = True

    if needs_rebuild:
        try:
            existing = existing.reindex(columns=REPORT_COLUMNS)
        except Exception:
            existing = pd.DataFrame(columns=REPORT_COLUMNS)
        _build_template(report_file, existing)


def resume_checks(processed_path: str, report_path: str, update_status=None, logger=None):
    """
    Reprend un processus de vérification interrompu.

    Cette fonction compare les partenaires dans le fichier de données traitées avec ceux
    déjà présents dans un fichier de rapport existant. Elle génère ensuite un rapport
    uniquement pour les partenaires manquants et les ajoute au rapport existant.

    Args:
        processed_path (str): Chemin vers le fichier de données traitées (`latest_datas.xlsx`).
        report_path (str): Chemin vers le fichier de rapport existant (`latest_report.xlsx`).
        update_status (callable, optional): Fonction de rappel pour les mises à jour de statut.
    """
    if update_status is None:
        update_status = lambda msg: None

    def _log_debug(msg: str):
        if logger is None:
            return
        if hasattr(logger, "debug"):
            logger.debug(msg)
        elif hasattr(logger, "log"):
            logger.log(f"[DEBUG] {msg}")

    update_status("Chargement des données...")
    _log_debug(f"Resume requested. Data={processed_path} Report={report_path}")

    try:
        df = pd.read_excel(processed_path, dtype=str).astype(str)
        df = _coerce_id_columns(df)
    except Exception as exc:
        raise RuntimeError(f"Impossible de lire le fichier data: {exc}")

    try:
        existing = pd.read_excel(report_path, sheet_name="Report", dtype=str).astype(str)
        existing = _coerce_id_columns(existing)
    except Exception:
        existing = pd.DataFrame(columns=REPORT_COLUMNS)

    # Identifier les partenaires qui sont dans le fichier de données mais pas encore dans le rapport.
    partner_col = "BP" if "BP" in df.columns else df.columns[0]
    done_partners = set(existing.get("BP", []))
    remaining = df[~df[partner_col].isin(done_partners)]

    _log_debug(f"Existing report has {len(existing)} lignes. Remaining partners to process: {len(remaining)}")
    if remaining.empty:
        update_status("Tout est déjà traité dans le rapport.")
        return report_path

    input_dir = os.path.dirname(report_path) or "."
    # Ensure template and existing data are in place before appending.
    _ensure_template_v2(report_path)

    # Keep column names consistent with report expectations.
    remaining = remaining.rename(columns={partner_col: "BP"})
    # Append only the missing rows.
    return generate_report(
        remaining,
        input_dir,
        output_dir=input_dir,
        update_status=update_status,
        report_path=report_path,
        logger=logger,
    )


def write_line(writer, report_df, line_df, save_path=None, sheet_name="Report"):
    """
    Ajoute une seule ligne de données à la feuille de rapport Excel et sauvegarde immédiatement le fichier.
    Cette approche "transactionnelle" empêche la perte de données si le script est arrêté.

    Args:
        writer (pd.ExcelWriter): L'objet ExcelWriter à utiliser.
        report_df (pd.DataFrame): Le DataFrame en mémoire contenant toutes les lignes traitées jusqu'à présent.
        line_df (pd.DataFrame): Un DataFrame d'une seule ligne contenant les nouvelles données à ajouter.
        save_path (str, optional): Le chemin explicite pour sauvegarder le fichier.
        sheet_name (str): Le nom de la feuille de calcul à laquelle ajouter.

    Returns:
        pd.DataFrame: Le DataFrame de rapport mis à jour.
    """
    # Accept both DataFrame and Series inputs; normalize to a one-row DataFrame.
    if isinstance(line_df, pd.Series):
        line_df = line_df.to_frame().T
    line_df = reports_col(line_df.astype(str))
    ordered = line_df.reindex(columns=REPORT_COLUMNS)
    report_df = pd.concat([report_df, ordered], ignore_index=True)
    start_row = writer.sheets[sheet_name].max_row
    ordered.to_excel(
        writer,
        index=False,
        header=False,
        startrow=start_row,
        sheet_name=sheet_name,
    )
    # Persist to disk after each append so a forced stop does not lose already processed lines.
    try:
        target = (
            save_path
            or getattr(writer, "path", None)
            or getattr(getattr(writer, "_handles", None), "handle", None)
        )
        if target is not None:
            writer.book.save(getattr(target, "name", target))
    except Exception:
        # Best-effort; do not break processing if the save hook fails.
        pass
    return report_df


def _get_status(line_df: pd.DataFrame) -> str:
    """
    Extrait en toute sécurité la valeur de statut d'un DataFrame d'une seule ligne.
    Renvoie une chaîne vide en cas d'erreur ou si la valeur est manquante.
    """
    try:
        return str(line_df.get("status", pd.Series([""])).iloc[0])
    except Exception:
        return ""


def _siren_only(row, BP, siren):
    """
    Gère la logique de vérification pour un SIREN uniquement.
    Appelle l'API SIRENE et formate la sortie dans un DataFrame d'une seule ligne.

    Args:
        row (pd.Series): La ligne de données d'origine du partenaire.
        BP (str): L'identifiant du partenaire.
        siren (str): Le numéro SIREN à vérifier.
    Returns:
        pd.DataFrame: Un DataFrame d'une seule ligne avec les résultats de la vérification.
    """
    if not (siren in ("", "None") or "Invalid input" in siren):
        infos_siren = rsn.handlesiren(siren)
        siren_line = pd.DataFrame([infos_siren]).astype(str)
        siren_line["type"] = "siren"
        api_cols = set(siren_line.columns)
        for key, value in row.items():
            if key == "siren":
                continue
            if key in api_cols:
                continue  # ne pas écraser les données renvoyées par l'API
            siren_line[key] = value
    else:
        siren_line = pd.DataFrame(
            [{"BP": BP, "type": "siren", "status": "invalid input"}]
        )
        for key, value in row.items():
            siren_line[key] = value
    return siren_line


def _siret_only(row, BP, siret):
    """
    Gère la logique de vérification pour un SIRET uniquement.
    Appelle l'API SIRENE et formate la sortie dans un DataFrame d'une seule ligne.

    Args:
        row (pd.Series): La ligne de données d'origine du partenaire.
        BP (str): L'identifiant du partenaire.
        siret (str): Le numéro SIRET à vérifier.
    Returns:
        pd.DataFrame: Un DataFrame d'une seule ligne avec les résultats de la vérification."""
    if not "Invalid input" in siret and siret not in ("", "None"):
        infos_siret = rst.handlesiret(siret)
        siret_line = pd.DataFrame([infos_siret]).astype(str)
        siret_line["type"] = "siret"
        api_cols = set(siret_line.columns)
        for key, value in row.items():
            if key in ("siret", "siren"):
                continue
            if key in api_cols:
                continue  # ne pas écraser les champs retournés par l'API
            siret_line[key] = value
    else:
        siret_line = pd.DataFrame(
            [{"BP": BP, "type": "siret", "status": "invalid input"}]
        )
        for key, value in row.items():
            siret_line[key] = value
    return siret_line

def reports_col(row):
    """
    Génère les colonnes de diagnostic et de rapport de synthèse.
    - `report`: Fournit un résumé lisible de l'état de la ligne.
    - `diagnostic_name`: Résultat de la comparaison de noms.
    - `diagnostic_street`: Résultat de la comparaison d'adresses.
    Args:
        row (pd.DataFrame): Le DataFrame (généralement une seule ligne) à traiter.
    Returns:        pd.DataFrame: Le DataFrame avec les nouvelles colonnes de diagnostic."""
    # Garantit la présence des colonnes de flags ajoutées récemment.
    for missing_col in [
        "missing siren",
        "missing siret",
        "Missing_Vat",
        "uses a snetor siren",
        "uses a snetor siret",
        "uses a snetor VAT",
        "Missmatching siren siret",
        "Missmatching siren VAT",
    ]:
        if missing_col not in row.columns:
            row[missing_col] = False

    cond_active = row["status"].isin(["Actif", "Active"])
    def _as_bool(series: pd.Series) -> pd.Series:
        return series.astype(str).str.lower().isin(["true", "1", "yes"])

    bad_flags = (
        _as_bool(row["missing siren"])
        | _as_bool(row["missing siret"])
        | _as_bool(row["Missing_Vat"])
        | _as_bool(row["uses a snetor siren"])
        | _as_bool(row["uses a snetor siret"])
        | _as_bool(row["uses a snetor VAT"])
        | _as_bool(row["Missmatching siren siret"])
        | _as_bool(row["Missmatching siren VAT"])
    )
    cond_all_good = ~bad_flags

    def _val_is_true(val) -> bool:
        return str(val).lower() in ("true", "1", "yes")

    def _flag_summary(row_idx):
        parts = []
        if _val_is_true(row.loc[row_idx, "missing siren"]):
            parts.append("missing siren")
        if _val_is_true(row.loc[row_idx, "missing siret"]):
            parts.append("missing siret")
        if _val_is_true(row.loc[row_idx, "Missing_Vat"]):
            parts.append("Missing_Vat")
        if _val_is_true(row.loc[row_idx, "uses a snetor siren"]):
            parts.append("uses a snetor siren")
        if _val_is_true(row.loc[row_idx, "uses a snetor siret"]):
            parts.append("uses a snetor siret")
        if _val_is_true(row.loc[row_idx, "uses a snetor VAT"]):
            parts.append("uses a snetor VAT")
        if _val_is_true(row.loc[row_idx, "Missmatching siren siret"]):
            parts.append("Missmatching siren siret")
        if _val_is_true(row.loc[row_idx, "Missmatching siren VAT"]):
            parts.append("Missmatching siren VAT")
        return ", ".join(parts) if parts else ""

    summaries = pd.Series({_idx: _flag_summary(_idx) for _idx in row.index})

    row.loc[~cond_all_good & cond_active, "report"] = summaries.loc[
        ~cond_all_good & cond_active
    ]
    row.loc[cond_all_good & ~cond_active, "report"] = "All good but siret/siren not active"
    row.loc[~cond_all_good & ~cond_active, "report"] = summaries.loc[
        ~cond_all_good & ~cond_active
    ]
    row.loc[cond_all_good & cond_active, "report"] = "Everything is valid"

    row["diagnostic_name"] = row.apply(compare_names, axis=1)
    return row

        

def generate_report(
    output: pd.DataFrame,
    input_dir: str,
    output_dir: str,
    update_status=None,
    report_path: Optional[str] = None,
    logger=None,
):
    """
    Génère le rapport de vérification final en itérant sur les données des partenaires,
    en appelant l'API SIRENE pour les SIREN/SIRET, et en écrivant les résultats dans un fichier Excel.

    La logique principale est de décider s'il faut vérifier le SIREN, le SIRET ou les deux,
    en fonction des données disponibles et de leur validité (par exemple, si le SIRET
    correspond au SIREN).

    Args:
        output (pd.DataFrame): Le DataFrame contenant les données des partenaires traitées.
        input_dir (str): Le répertoire où le rapport sera sauvegardé.
        update_status (callable, optional): Fonction de rappel pour les mises à jour de statut.
        report_path (Optional[str]): Chemin explicite vers le fichier de rapport. S'il est fourni,
                                     les nouvelles lignes seront ajoutées à ce fichier.
        logger (optional): Logger optionnel (interface .log/.info/.warn) pour tracer les statuts.
    """
    if update_status is None:
        update_status = lambda msg: None

    def _log_info(msg: str):
        if logger is None:
            return
        if hasattr(logger, "log"):
            logger.log(msg)
        elif hasattr(logger, "info"):
            logger.info(msg)
        else:
            try:
                logger(msg)
            except Exception:
                pass

    def _log_warn(msg: str):
        if logger is None:
            return
        if hasattr(logger, "warn"):
            logger.warn(msg)
        elif hasattr(logger, "warning"):
            logger.warning(msg)
        else:
            _log_info(f"[WARN] {msg}")

    def _log_debug(msg: str):
        if logger is None:
            return
        if hasattr(logger, "debug"):
            logger.debug(msg)
        else:
            _log_info(f"[DEBUG] {msg}")

    update_status("Generation du rapport (verification SIREN/SIRET)...")
    _log_debug(f"Report generation start: {len(output)} rows -> {output_dir}")

    report_file = report_path or os.path.join(output_dir, "latest_report.xlsx")
    update_status("Initialisation du rapport Excel...")
    # Recrée systématiquement le gabarit pour éviter tout décalage de colonnes.
    _build_template(report_file)

    # Normalise une fois la structure des colonnes pour le traitement en aval.
    output = output.reindex(columns=REPORT_COLUMNS, fill_value="").astype(str)
    report = pd.DataFrame(columns=REPORT_COLUMNS).astype(str)
    report = report.reindex(columns=REPORT_COLUMNS)
    n_out = len(output)

    # Ouvre le fichier Excel en mode "append" en utilisant openpyxl.
    # 'if_sheet_exists="overlay"' permet d'écrire sur une feuille existante.
    with pd.ExcelWriter(
        report_file,
        engine="openpyxl",
        mode="a",
        if_sheet_exists="overlay",
    ) as writer:
        # Walk through partners and decide which API calls to trigger (SIREN, SIRET, or both).
        for counter, (_, row) in enumerate(output.iterrows(), start=1):
            BP = row.get("BP")
            siren = str(row.get("siren"))
            siret = str(row.get("siret"))

            # Normalize Excel-like floats: "78415854500110.0" -> "78415854500110"
            # Nettoie les numéros qui pourraient être interprétés comme des flottants par Excel.
            if siret.endswith(".0"):
                siret = siret.split(".")[0]
            if siren.endswith(".0"):
                siren = siren.split(".")[0]

            update_status(f"Verification {counter}/{n_out} : partenaire {BP}")
            _log_debug(f"Checking {counter}/{n_out} BP={BP} siren={siren} siret={siret}")

            # Case 1: missing/invalid SIRET -> check SIREN only
            # Si le SIRET est manquant ou invalide, on ne vérifie que le SIREN.
            if siret.lower() in ("", "none") or "invalid input" in siret.lower() or siret.lower() == "nan":
                siren_line = _siren_only(row, BP, siren)
                report = write_line(writer, report, siren_line, report_file)
                _log_info(f"BP={BP} type=siren status={_get_status(siren_line)}")

            # Case 2: SIRET present but not starting with SIREN -> check both SIREN and SIRET
            # Si le SIREN et le SIRET ne correspondent pas, les deux sont vérifiés séparément.
            elif siret[:9] != siren and not (siren in ("", "None") or "Invalid input" in siren):
                siren_line = _siren_only(row, BP, siren)
                report = write_line(writer, report, siren_line, report_file)
                _log_info(f"BP={BP} type=siren status={_get_status(siren_line)}")

                siret_line = _siret_only(row, BP, siret)
                report = write_line(writer, report, siret_line, report_file)
                _log_info(f"BP={BP} type=siret status={_get_status(siret_line)}")

            # Case 3: SIRET present and starts with SIREN -> check SIRET only
            # Si le SIRET est présent et semble valide (commence par le SIREN), on ne vérifie que le SIRET.
            elif siret[:9] == siren and not (siret in ("", "None") or "Invalid input" in siret):
                siret_line = _siret_only(row, BP, siret)
                report = write_line(writer, report, siret_line, report_file)
                siret_status = _get_status(siret_line)
                _log_info(f"BP={BP} type=siret status={siret_status}")
                # Cas particulier : si le SIRET est inactif, on effectue quand même une vérification du SIREN
                # pour voir si l'unité légale elle-même est toujours active.
                if siret_status not in ("Active", "Actif") and isinstance(siren, str) and len(siren) == 9:
                    siren_line = _siren_only(row, BP, siren)
                    report = write_line(writer, report, siren_line, report_file)
                    _log_warn(f"BP={BP} inactive SIRET -> added SIREN check status={_get_status(siren_line)}")
            else:
                line = row
                report = write_line(writer, report, line, report_file)
                _log_warn(f"BP={BP} skipped: no valid SIREN/SIRET provided")

    return report_file


def main(
    output: pd.DataFrame,
    input_dir: str,
    output_dir: str,
    update_status=None,
    report_path: Optional[str] = None,
    logger=None,
    ):
    try:
        generate_report(
            output=output,
            input_dir=input_dir,
            output_dir=output_dir,
            update_status=update_status,
            report_path=report_path,
            logger=logger,
            )
    except Exception as exc:
        logger.error(f"Unexpected error in SIREN/SIRET verification pipeline: \n{exc}")
        raise exc